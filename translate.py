'''
- Takes ingredients list as .xlsx and outputs as .docx
- Example spreadsheet "sample_data.xlsx" is provided
- Spreadsheet data is grouped into fields (columns):
    - unique_id, sorting_id1, sorting_id2, and three different locales
'''
import openpyxl
from docx import Document

# get workbook and sheet
wb = openpyxl.load_workbook( 'sample_data.xlsx' )
sheet = wb.get_sheet_by_name( 'main' )

# unique_id is type 'int', column 1
# sorting_id1 is type 'float', column 2
# all other fields are type 'str', column 3, 4, 5
# print( sheet.cell( row = 1, column = 1 ).value )
# print( type( sheet.cell( row = 2, column = 1 ).value ) )

# field/col indices
col_sorting_id1 = 0 # first value in tuple
col_sorting_id2 = 1 # second value etc.
col_locale_en_US = 2
col_locale_fr_FR = 3
col_locale_es_ES = 4

def parse_sheet( sheet ):
    '''
    - parse worksheet, get values as dict of tuples
    - must tailor this function to your own spreadsheet's format
    '''
    # row indices
    row = 2 # parsing starts from this row
    max_row = sheet.max_row # last populated row in sheet

    # field/col indices: 1 = A, 2 = B, etc.
    parse_col_key = 1 # dict keys always in this column
    parse_col_first = 2 # first value in tuple will be this column
    parse_col_second = 3 # second value etc.
    parse_col_third = 4
    parse_col_fourth = 5
    parse_col_fifth = 6

    table = {}
    for parse_col_key in range( parse_col_key, max_row ):
        table[parse_col_key] = (
            sheet.cell( row = row, column = parse_col_first ).value,
            sheet.cell( row = row, column = parse_col_second ).value,
            sheet.cell( row = row, column = parse_col_third ).value,
            sheet.cell( row = row, column = parse_col_fourth ).value,
            sheet.cell( row = row, column = parse_col_fifth ).value
            )
        row += 1

    return table

def sort_by_field( field ):
    '''
    - returns a sorted list of ALL the values in the spreadsheet
    - can sort by different fields
    '''
    table = parse_sheet( sheet )
    table_sorted = sorted( table.values(),
                            key = lambda data: data[field] )

    return table_sorted

def get_field( target_field, sorting_field = col_locale_en_US ):
    table = parse_sheet( sheet )
    table_sorted = sort_by_field( sorting_field )

    retrieved = [
        table_sorted[table_sorted.index( n )][target_field]
        for n in table_sorted
    ]

    return retrieved

def delimit( field, language ):
    if language == 'en':
        out = 'INGREDIENTS: '
    elif language == 'es':
        out = 'INGREDIENTES: '
    elif language == 'fr':
        out = 'INGREDIENTS: '
    for item in field:
        out += str( item ) + ', '

    out = out[:-2]
    out += '.'
    return out

def write_to_docx( text, name ):
    doc = Document()
    par = doc.add_paragraph( text )
    doc.save( name + '.docx' )

'''
USER PROMPTS
'''
print( '----------------------------')
print( '| Excel to Word Translator |' )
print( '----------------------------')
print( 'Parses and sorts an .xlsx file and outputs contents as a .docx file as comma-delimited paragraph.\n' )
print( '- .xlsx file MUST be in the same directory as this script.' )
print( '- .docx file will be output in the same directory as this script.\n' )
print( 'Press CTRL+C to cancel this script.\n' )
print( '#########################################################\n' )
# get values in locale_en_US, sort by same field
target = False
while not target:
    target_field = input( "Enter the field to retrieve.\n\tsorting_id1 = 1\n\tsorting_id2 = 2\n\tlocale_en_US = 3\n\tlocale_fr_FR = 4\n\tlocale_es_ES = 5\n\t(To retrieve sorting_id1, enter 1)\n>> " )

    if target_field == '1':
        target_field = col_sorting_id1
        target = True
    elif target_field == '2':
        target_field = col_sorting_id2
        target = True
    elif target_field == '3':
        target_field = col_locale_en_US
        target = True
    elif target_field == '4':
        target_field = col_locale_fr_FR
        target = True
    elif target_field == '5':
        target_field = col_locale_es_ES
        target = True
    else:
        print( '\nSorry, don\'t recognize that number.' )

sort = False
while not sort:
    sorting_field = input( "\nEnter the field to sort by.\n\tsorting_id1 = 1\n\tsorting_id2 = 2\n\tlocale_en_US = 3\n\tlocale_fr_FR = 4\n\tlocale_es_ES = 5\n\t(To sort by sorting_id1, enter 1)\n>> " )

    if sorting_field == '1':
        sorting_field = col_sorting_id1
        sort = True
    elif sorting_field == '2':
        sorting_field = col_sorting_id2
        sort = True
    elif sorting_field == '3':
        sorting_field = col_locale_en_US
        sort = True
    elif sorting_field == '4':
        sorting_field = col_locale_fr_FR
        sort = True
    elif sorting_field == '5':
        sorting_field = col_locale_es_ES
        sort = True
    else:
        print( '\nSorry, don\'t recognize that number.' )

lang = False
while not lang:
    lang_out = input( "\nEnter the display language.\n\tEnglish = 1\n\tFrench = 2\n\tSpanish = 3\n>> " )

    if lang_out == '1':
        lang_out = 'en'
        lang = True
    elif lang_out == '2':
        lang_out = 'fr'
        lang = True
    elif lang_out == '3':
        lang_out = 'es'
        lang = True


doc_name = input( "\nEnter the output file name WITHOUT the file extension.\n\t(Entering \"test\" will save a file called \"test.docx\").\n>> ")

values = get_field( target_field, sorting_field )
field_delimited = delimit( values, lang_out )
write_to_docx( field_delimited, doc_name )
